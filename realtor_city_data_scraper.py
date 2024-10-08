import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException
import time
import random
import pandas as pd
from pathlib import Path
from datetime import datetime
from tqdm import tqdm
from openpyxl import Workbook


# Base URL for scraping
baseURL = "https://www.realtor.ca/map#view=list&CurrentPage={}&Sort=6-D&GeoIds=g30_dpz89rm7&GeoName={}%2C%20ON&PropertyTypeGroupID=1&TransactionTypeId=2&PropertySearchTypeId=1&Currency=CAD&HiddenListingIds=&IncludeHiddenListings=false"


def setup_chrome_options():
    chrome_options = uc.ChromeOptions()
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    # chrome_options.add_argument("--headless--")
    return chrome_options


def setup_driver():
    chrome_options = setup_chrome_options()
    driver = uc.Chrome(options=chrome_options)
    driver.maximize_window()
    wait = WebDriverWait(driver, 15)
    return driver, wait


def check_for_incapsula_captcha(driver):
    page_source = driver.page_source
    return "Request unsuccessful. Incapsula incident ID:" in page_source


def get_and_save_property_links(driver, city_name, page_count=50):
    page_links = {}
    filename = f"property_links_{city_name}.txt"

    for page_number in range(1, page_count + 1):
        url = baseURL.format(page_number, city_name)
        print(f"Getting URL from page#{page_number} for {city_name}")
        
        driver.get(url)
        driver.refresh()

        # wait for the page to load
        WebDriverWait(driver, 15).until(
            EC.presence_of_all_elements_located((By.CLASS_NAME, 'cardCon'))
        )

        time.sleep(3)
        print(f"Got URL from page#{page_number} for {city_name}\n")

        # scroll and fetch property links
        driver.execute_script(f"window.scrollBy(0, {1050});")
        time.sleep(1)

        li_elements = driver.find_elements(By.CLASS_NAME, 'cardCon')
        links = [li.find_element(By.CLASS_NAME, 'listingDetailsLink').get_attribute('href') for li in li_elements]
        
        page_links[f"Page {page_number}"] = links
        time.sleep(5)
    
    # write the links to the file
    with open(filename, 'w') as f:
        for page, links in page_links.items():
            f.write(f"{page}:\n")
            for link in links:
                f.write(f"{link}\n")
            f.write("\n")
    
    print(f"Property links for {city_name} saved to {filename}")
    return page_links


def scrape_property_data(driver, link):
    scraped_data = {}
    try:
        driver.get(link)
        time.sleep(5)

        while check_for_incapsula_captcha(driver):
            print("Access blocked by Incapsula. Please solve the CAPTCHA and type 'solved' to continue.")
            input_text = input("Enter 'solved' after resolving the issue: ")
            if input_text.strip().lower() == 'solved':
                break

        scroll_down_amount = random.randint(300, 1000)
        driver.execute_script(f"window.scrollBy(0, {scroll_down_amount});")

        # Extract property details
        price = driver.find_element(By.XPATH, "//div[@id='listingPriceValue']").text.upper()
        address_full = driver.find_element(By.XPATH, "//h1[@id='listingAddress']").text.upper()
        address_lines = address_full.split("\n")
        address = address_lines[0]
        city_postal = address_lines[1].split(",")
        city = city_postal[0].strip().split(" (")[0]
        state_postal = city_postal[1].strip().split()
        state = state_postal[0].strip()
        postal_code = state_postal[1] if len(state_postal) > 1 else ''
        agent_name = driver.find_element(By.XPATH, "//span[@class='realtorCardName']").text.upper()
        broker_name = driver.find_element(By.XPATH, "//div[@class='officeCardName']").text.upper()

        # Get lat and long
        directions_btn = driver.find_element(By.ID, 'listingDirectionsBtn')
        directions_href = directions_btn.get_attribute('href')
        lat_long = directions_href.split("destination=")[-1].replace('%2c', ',')
        latitude, longitude = lat_long.split(",")

        # Store the scraped data
        scraped_data = {
            "Address": address,
            "City": city,
            "State": state,
            "Postal Code": postal_code,
            "Agent": agent_name,
            "Broker": broker_name,
            "Price": price,
            "Latitude": latitude.strip(),
            "Longitude": longitude.strip()
        }

        scroll_up_amount = random.randint(300, 1000)
        driver.execute_script(f"window.scrollBy(0, {scroll_up_amount});")
        time.sleep(5)

    except Exception as e:
        print(f"Failed to scrape {link}: {e}")

    return scraped_data


def generate_excel_filename(city, prefix="GM1"):
    today = datetime.today()
    formatted_file_name = f"{prefix}{today.month:02d}{today.day:02d}_{city.capitalize()}.xlsx"
    return formatted_file_name


def save_data_to_excel(scraped_data_list, city):
    file_name = generate_excel_filename(city)
    df = pd.DataFrame(scraped_data_list)
    df.drop_duplicates()
    # df.drop_duplicates(subset=['Address', 'Agent', 'Broker', 'Price'])
    df.to_excel(file_name, index=False)
    print(f"Data saved to {file_name}")


def create_workbook_by_postal_code(scraped_data_excel_filepath):
    df = pd.read_excel(scraped_data_excel_filepath)
    wb = Workbook()
    for prefix in range(1, 10):
        mask = df['Postal Code'].str.startswith(f'M{prefix}')
        filtered_df = df[mask]
        if not filtered_df.empty:
            sheet = wb.create_sheet(title=f'M{prefix}')
            sheet.append(filtered_df.columns.tolist())
            for row in filtered_df.itertuples(index=False):
                sheet.append(row)
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(Path.stem(scraped_data_excel_filepath))


def scrape_all_property_data(driver, links_file):
    scraped_data_list = []
    with open(links_file, 'r') as file:
        lines = file.readlines()

    for line in tqdm(lines, desc="Scraping property data", unit="link", ncols=100):
        link = line.strip()
        if link.startswith("http"):
            data = scrape_property_data(driver, link)
            if data:
                scraped_data_list.append(data)

    return scraped_data_list


# Main execution
if __name__ == "__main__":
    cities = ["Toronto", "Mississauga"]
    driver, wait = setup_driver()
    
    for city in cities:
        _ = get_and_save_property_links(driver, city)


    driver.quit()